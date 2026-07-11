import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DATABASE = "data/database/nifty100.db"

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE"
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C"
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE"
)

GOLD_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966"
)

class PeerComparisonReport:

    def __init__(self):

        self.conn = sqlite3.connect(DATABASE)

        print("Peer Comparison Report Loaded")

        self.peer_df = pd.read_sql(
            "SELECT * FROM peer_groups",
            self.conn
        )

        self.percentile_df = pd.read_sql(
            "SELECT * FROM peer_percentiles",
            self.conn
        )

        self.ratio_df = pd.read_sql(
            "SELECT * FROM financial_ratios",
            self.conn
        )

        self.company_df = pd.read_sql(
            "SELECT * FROM companies",
            self.conn
        )

        print(self.company_df.columns.tolist())

        self.ratio_df = (
            self.ratio_df
            .sort_values("id")
            .drop_duplicates(
                subset=["company_id", "year"],
                keep="first"
            )
        )

    def get_peer_groups(self):

        return sorted(
            self.peer_df["peer_group_name"].unique()
        )

    def get_peer_report(self, peer_group):

        companies = self.peer_df[
            self.peer_df["peer_group_name"] == peer_group
        ]

        companies = companies.merge(
            self.company_df[
                ["id", "company_name"]
            ],
            left_on="company_id",
            right_on="id",
            how="left"
        )

        companies = companies.drop(
            columns=["id_x", "id_y"]
        )

        return companies
    
    def get_financial_data(self, peer_group, year):

        companies = self.get_peer_report(peer_group)

        financials = self.ratio_df[
            self.ratio_df["year"] == year
        ]

        report = companies.merge(
            financials,
            on="company_id",
            how="left"
        )

        return report
    
    def get_percentile_data(self, peer_group, year):

        report = self.get_financial_data(
            peer_group,
            year
        )

        percentiles = self.percentile_df[
            (self.percentile_df["peer_group_name"] == peer_group) &
            (self.percentile_df["year"] == year)
        ]

        return report, percentiles
    
    def pivot_percentiles(self, percentile_df):

        pivot = percentile_df.pivot(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        )

        pivot = pivot.reset_index()

        return pivot
    
    def pivot_percentiles(self, percentile_df):

        pivot = percentile_df.pivot(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        ).reset_index()

        # Rename percentile columns
        pivot = pivot.rename(
            columns={
                col: f"{col}_percentile"
                for col in pivot.columns
                if col != "company_id"
            }
        )

        return pivot
    
    def build_peer_report(self, peer_group, year):

        financials, percentiles = self.get_percentile_data(
            peer_group,
            year
        )

        pivot = self.pivot_percentiles(percentiles)

        report = financials.merge(
            pivot,
            on="company_id",
            how="left"
        )

        return report
    
    def export_peer_report(self, year):

        peer_groups = self.get_peer_groups()

        with pd.ExcelWriter(
            "reports/peer_comparison.xlsx",
            engine="openpyxl"
        ) as writer:

            for peer_group in peer_groups:

                print(f"Exporting {peer_group}")

                report = self.build_peer_report(
                    peer_group,
                    year
                )

                report.to_excel(
                    writer,
                    sheet_name=peer_group,
                    index=False
                )

        self.format_percentiles()

        self.add_median_row()

        print("Peer comparison report created successfully.")      

    def format_percentiles(self):

        wb = load_workbook("reports/peer_comparison.xlsx")

        for sheet in wb.worksheets:

            print(f"Formatting {sheet.title}")

            headers = [
                cell.value
                for cell in sheet[1]
            ]

            # Find all percentile columns
            percentile_columns = []

            for i, header in enumerate(headers, start=1):

                if (
                    header is not None
                    and str(header).endswith("_percentile")
                ):
                    percentile_columns.append(i)

            # Find benchmark column
            benchmark_col = headers.index("is_benchmark") + 1

            # Highlight benchmark company row
            for row in range(2, sheet.max_row + 1):

                benchmark = sheet.cell(
                    row=row,
                    column=benchmark_col
                ).value

                if benchmark == 1:

                    for col in range(1, sheet.max_column + 1):

                        sheet.cell(
                            row=row,
                            column=col
                        ).fill = GOLD_FILL

            # Color percentile columns
            for col in percentile_columns:

                for row in range(2, sheet.max_row + 1):

                    cell = sheet.cell(
                        row=row,
                        column=col
                    )

                    if cell.value is None:
                        continue

                    if cell.value >= 75:

                        cell.fill = GREEN_FILL

                    elif cell.value <= 25:

                        cell.fill = RED_FILL

                    else:

                        cell.fill = YELLOW_FILL

        wb.save("reports/peer_comparison.xlsx")

        print("Percentile formatting complete.")

    def add_median_row(self):

        wb = load_workbook("reports/peer_comparison.xlsx")

        for sheet in wb.worksheets:

            print(f"Adding median row to {sheet.title}")

            last_row = sheet.max_row + 1

            # Label
            sheet.cell(row=last_row, column=1).value = "Median"

            headers = [
                cell.value
                for cell in sheet[1]
            ]

            for col in range(2, sheet.max_column + 1):

                header = headers[col - 1]

                # Skip text columns
                if header in [
                    "company_name",
                    "company_id",
                    "peer_group_name"
                ]:
                    continue

                values = []

                for row in range(2, last_row):

                    value = sheet.cell(
                        row=row,
                        column=col
                    ).value

                    if isinstance(value, (int, float)):
                        values.append(value)

                if values:

                    sheet.cell(
                        row=last_row,
                        column=col
                    ).value = round(pd.Series(values).median(), 2)

        wb.save("reports/peer_comparison.xlsx")

        print("Median rows added successfully.")